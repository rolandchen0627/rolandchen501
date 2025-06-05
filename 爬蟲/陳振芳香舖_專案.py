import os
import re
import time
import requests
import pandas as pd
from bs4 import BeautifulSoup
from datetime import datetime

from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter


def scrape_jobs():
    """開始爬取產品資訊並下載圖片"""
    job_data = []
    keyword = "香道用品"
    today_date = datetime.today().strftime('%Y-%m-%d')
    output_folder = r'C:\Users\\Desktop\爬蟲\芳香專案'
    image_folder = os.path.join(output_folder, "images")
    os.makedirs(output_folder, exist_ok=True)
    os.makedirs(image_folder, exist_ok=True)
    output_file = os.path.join(output_folder, f'陳振芳香舖_{today_date}_{keyword}.xlsx')

    url = 'https://www.unikgn.com.tw/cht/index.php?code=list&ids=19&type_id=8'
    res = requests.get(url)
    res.raise_for_status()

    soup = BeautifulSoup(res.text, 'html.parser')
    articles = soup.find_all('div', class_="pro-list-item mb-4")

    for idx, article in enumerate(articles, 1):
        name_tag = article.find('h1', class_="h5")
        product_name = name_tag.text.strip() if name_tag else "未知"

        price_tag = article.find('div', class_='price-content')
        product_price = price_tag.text.strip() if price_tag else "未知"

        image_div = article.find('div', class_='pro-img')
        style = image_div.get('style', '')
        match = re.search(r"url\((.*?)\)", style)
        product_img_url = "https://www.unikgn.com.tw" + match.group(1).lstrip(".") if match else ""

        # 下載圖片
        image_filename = "無圖"
        if product_img_url:
            try:
                sanitized_name = re.sub(r'[\\/:"*?<>|]', '_', product_name)  # 清除非法檔名字元
                image_filename = os.path.join(image_folder, f'{idx:02d}_{sanitized_name}.jpg')
                img_res = requests.get(product_img_url)
                img_res.raise_for_status()
                with open(image_filename, 'wb') as f:
                    f.write(img_res.content)
            except Exception as e:
                print(f"⚠️ 無法下載圖片 {product_img_url}：{e}")
                image_filename = "下載失敗"

        job_info = {
            "產品種類": keyword,
            "產品名稱": product_name,
            "價格": product_price,
            "本地圖片路徑": image_filename
        }
        job_data.append(job_info)

        print(f"📌 {product_name} | {product_price} | 已儲存圖片: {image_filename}")

    return job_data, output_file


def save_to_excel_with_images(job_data, output_file):
    """儲存產品資訊到 Excel 並插入圖片"""
    wb = Workbook()
    ws = wb.active
    ws.title = "產品清單"

    headers = ["產品種類", "產品名稱", "價格", "圖片"]
    ws.append(headers)

    for idx, item in enumerate(job_data, start=2):  # 從第2列開始寫入資料
        ws.cell(row=idx, column=1, value=item["產品種類"])
        ws.cell(row=idx, column=2, value=item["產品名稱"])
        ws.cell(row=idx, column=3, value=item["價格"])

        # 插入圖片
        img_path = item.get("本地圖片路徑")
        if os.path.exists(img_path) and img_path.lower().endswith((".jpg", ".png")):
            try:
                img = OpenpyxlImage(img_path)
                img.width = 80  # 圖片寬
                img.height = 80  # 圖片高
                ws.row_dimensions[idx].height = 60  # Excel 行高，讓圖片顯示不被裁切
                ws.add_image(img, f"D{idx}")  # 插入圖片到 E 欄
            except Exception as e:
                print(f"⚠️ 插入圖片失敗 {img_path}：{e}")

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 25

    wb.save(output_file)


def main():
    job_data, output_file = scrape_jobs()
    save_to_excel_with_images(job_data, output_file)


if __name__ == "__main__":
    main()
