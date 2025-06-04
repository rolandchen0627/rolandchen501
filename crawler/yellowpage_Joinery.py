from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import time
import openpyxl
from openpyxl import Workbook
import pandas as pd
import datetime
import os

keyword = "Joinery"
start_time = time.time()
current_date = datetime.datetime.now().strftime('%Y%m%d')
# 瀏覽器設定檔
option = webdriver.ChromeOptions()
prefs = {"profile.default_content_setting_values": {"notifications": 2}}
option.add_experimental_option("prefs", prefs)

driver = webdriver.Chrome(service=Service(
    executable_path="C:\\Users\\Desktop\\chromedriver.exe"), options=option)
driver.get("https://www.yellowpages-uae.com/")


WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
    (By.CSS_SELECTOR, 'input[placeholder="Products, Services, Brand or Company"]'))).send_keys(keyword)
WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
    (By.CSS_SELECTOR, 'button[class="p-1 px-2 bg-gray-800 font-semibold text-white rounded-[2.5px] w-full max-w-[90px]"]'))).click()

excel_file = f"C:/Users/Desktop/爬蟲/UAE/{keyword}.xlsx"

try:
    wb = openpyxl.load_workbook(excel_file)
except FileNotFoundError:
    wb = Workbook()
    wb.save(excel_file) 
    ws = wb.active

wb = openpyxl.load_workbook(excel_file)
ws = wb.active

ws.title = f'{keyword}'
ws['A1'] = 'Company_name'
ws['B1'] = 'Location'
ws['C1'] = 'city'
ws['D1'] = 'P.O BOX'
ws['E1'] = 'Phone'
ws['F1'] = 'Mobile'

n = 1

# 爬蟲邏輯
while True:
    soup = BeautifulSoup(driver.page_source, "html.parser")

    
    divs = soup.find_all("div")
    for div in divs:
        
        div_classes = div.get("class", [])

        # "row" 和 "box" 的 div
        if "row" in div_classes and "box" in div_classes:
            n += 1

            # 提取公司名稱
            company_name_tag = div.find("span")
            if company_name_tag:
                company_name = company_name_tag.text.strip()
            else:
                company_name = "Company name not found"
            ws[f'A{n}'] = company_name

            # 提取 Location 資訊
            location_tag = div.find("span", class_="font-semibold", text="Location : ")
            if location_tag:
                location = location_tag.find_next("span").text.strip() 
            else:
                location = "Location not found"
            ws[f'B{n}'] = location

            # 提取 City 資訊
            city_tag = div.find("span", class_="font-semibold", text="City : ")
            if city_tag:
                city = city_tag.find_next("span").text.strip()
            else:
                city = "City not found"
            ws[f'C{n}'] = city

            # 提取 P.O Box 資訊
            box_tag = div.find("span", class_="font-semibold", text="P.O Box : ")
            if box_tag:
                box = box_tag.find_next("span").text.strip()
            else:
                box = "P.O Box not found"
            ws[f'D{n}'] = box

            # 提取 Phone 資訊
            phone_tag = div.find("a", attrs={"aria-label": "Phone", "title": "Click to call"})
            if phone_tag:
                phone = phone_tag.text.strip()
            else:
                phone = "Phone not found"
            ws[f'E{n}'] = phone

            # 提取 Mobile 資訊
            mobile_tag = div.find("a", attrs={"aria-label": "Mobile", "title": "Click to call"})
            if mobile_tag:
                mobile = mobile_tag.text.strip()
            else:
                mobile = "Mobile not found"
            ws[f'F{n}'] = mobile

            # 打印公司名稱確認
            print(f'{n-1}..{company_name}')

    # 檢查是否存在下一頁按鈕
    next_button = soup.find("button", class_="border-[1px] border-gray-400 rounded-lg px-2 h-[30px]", text="Next")
    if not next_button or next_button.get('disabled') or next_button.get('value') == 'false':
        print("No more pages or 'next page' button is disabled.")
        break

    # 點擊下一頁按鈕
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
        (By.XPATH, '//*[text()="Next"]'))).click()
    time.sleep(3)

wb.save(excel_file)
end_time = time.time()
execution_time = end_time - start_time
hours, rem = divmod(execution_time, 3600)
minutes, seconds = divmod(rem, 60)
save_path = r'Z:\爬蟲名單\加盟商\UAE'
a = pd.read_excel(excel_file)
data = a.drop_duplicates(subset=['Company_name'], keep='first', inplace=False)

file_name = f'{current_date}_{keyword}_deduplication.xlsx'
full_path = os.path.join(save_path, file_name)
data.to_excel(full_path, index=False)


print(f'檔案已儲存至: {full_path}')
print(f'程式執行時間: {int(hours)} 小時 {int(minutes)} 分 {seconds:.2f} 秒')


driver.close()