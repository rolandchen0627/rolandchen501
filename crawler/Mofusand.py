# === Standard Libraries ===
import os
import time
from datetime import datetime
from io import BytesIO

# === Third-Party Libraries ===
import requests
import pandas as pd
from PIL import Image
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage

# === Selenium ===
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options


##提取網址###
def scrape_product_urls_selenium():
    base_url = 'https://www.miravivi.com.tw/v2/official/SalePageCategory/489616?sortMode=Curator'
    base_url_prefix = 'https://www.miravivi.com.tw'

    # 設定 Selenium（無頭模式）
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    service = Service()  # 自動使用預設安裝好的 ChromeDriver

    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.get(base_url)

    time.sleep(5)  # 等待 JavaScript 載入

    # 自動滾動頁面到底（模擬 lazy loading）
    last_height = driver.execute_script("return document.body.scrollHeight")
    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height

    # 抓取所有 href
    a_tags = driver.find_elements(By.TAG_NAME, "a")
    urls = []
    for a in a_tags:
        href = a.get_attribute('href')
        if href and '/SalePage/Index/' in href:
            urls.append(href)

    driver.quit()

    # 去重並儲存
    urls = sorted(set(urls))

    today = datetime.today().strftime('%Y-%m-%d')
    output_folder = r'C:\Users\Desktop\爬蟲\Mofusand'
    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, f'Mofusand商品網址_{today}.xlsx')

    df = pd.DataFrame(urls, columns=['商品網址'])
    df.to_excel(output_file, index=False)

if __name__ == "__main__":
    scrape_product_urls_selenium()
    


# 抓取產品資訊
def scrape_product_info(url):
    options = Options()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    driver = webdriver.Chrome(options=options)
    driver.get(url)
    time.sleep(3)

    # 抓產品名稱
    try:
        name = driver.find_element(By.CLASS_NAME, "salepage-title").text.strip()
    except:
        name = "找不到產品名稱"

    # 抓價格
    try:
        price_box = driver.find_element(By.CLASS_NAME, "salepage-price")
        price = price_box.find_element(By.TAG_NAME, "span").text.strip()
    except:
        price = "找不到價格"

    # 抓第二張圖片網址
    try:
        img_element = driver.find_element(By.ID, "ns-media-gallery-main-img")
        img_url = img_element.get_attribute("src").replace("/0/", "/1/")
    except:
        img_url = "找不到圖片"

    driver.quit()
    return name, price, img_url

# Excel 檔案路徑
input_path = r"C:\Users\Desktop\爬蟲\Mofusand\Mofusand商品網址.xlsx"
output_path = r"C:\Users\Desktop\爬蟲\Mofusand\mofusand_商品資料.xlsx"

# 讀取原始 Excel
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# 建立新 Excel
new_wb = openpyxl.Workbook()
new_ws = new_wb.active
new_ws.title = "Mofusand商品"
new_ws.append(["產品名稱", "價格", "圖片網址"])

# 從 A2 開始讀取網址
for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
    url = row[0].value
    if not url:
        continue
    name, price, img_url = scrape_product_info(url)
    new_ws.append([name, price, img_url])

# 儲存新 Excel
new_wb.save(output_path)



# Excel 路徑
excel_path = r"C:\Users\11020964.TPTWKD\Desktop\Github\爬蟲\Mofusand\mofusand_商品資料.xlsx"
output_folder = os.path.dirname(excel_path)

# 載入 Excel
wb = load_workbook(excel_path)
ws = wb.active

# 設定目標儲存格的欄寬與列高
ws.column_dimensions["D"].width = 15  # 調整 D 欄寬度
default_row_height = 80  # 調整列高度（要搭配圖片大小）

# 從 C2 開始
row = 2
while True:
    url = ws[f"C{row}"].value
    if not url:
        break

    try:
        # 下載圖片
        response = requests.get(url, timeout=10)
        response.raise_for_status()

        image_name = f"mofusand_image_{row}.jpg"
        image_path = os.path.join(output_folder, image_name)

        # 轉為 PIL 圖片並縮圖
        img = Image.open(BytesIO(response.content))
        img.thumbnail((100, 100))  # 控制圖片大小（儲存格內）
        img.save(image_path)

        # 插入圖片
        excel_img = ExcelImage(image_path)
        excel_img.width, excel_img.height = img.size
        ws.row_dimensions[row].height = default_row_height  # 設定列高
        ws.add_image(excel_img, f"D{row}")

        print(f"✅ 圖片插入 D{row}")

    except Exception as e:
        print(f"❌ 錯誤於 C{row}: {e}")

    row += 1

# 儲存 Excel
wb.save(excel_path)
print("✅ Excel 儲存完成")