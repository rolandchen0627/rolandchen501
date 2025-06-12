from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import time
import openpyxl
import pandas as pd
from openpyxl import Workbook
from selenium.webdriver.common.keys import Keys
import requests

# 瀏覽器設定
option = webdriver.ChromeOptions()
prefs = {"profile.default_content_setting_values": {"notifications": 2}}
option.add_experimental_option("prefs", prefs)

driver = webdriver.Chrome(service=Service(
    executable_path="C:\\Users\\11020964\\Desktop\\python\\chromedriver.exe"), options=option)
base_url = 'https://www.zycg.gov.cn/freecms/site/zygjjgzfcgzx/ddwtxm/index.html'
driver.get(base_url)


WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.CSS_SELECTOR, 'li[names="中标（成交）公告"]'))
).click()

excel_file = "C:/Users/11020964/Desktop/py/爬蟲/大陸/中央政府採購網0113.xlsx"

try:
    wb = openpyxl.load_workbook(excel_file)
except FileNotFoundError:
    wb = Workbook()  
    wb.save(excel_file)  
    ws = wb.active

wb = openpyxl.load_workbook(excel_file)
ws = wb.active

ws.title = '中央政府採購網'
ws['A1'] = '標題'
ws['B1'] = '發佈時間'
ws['C1'] = 'Url'
n = 2  

href_get = []  
max_per_page = 12
count_per_page = 0

while True:
    element_present = EC.presence_of_element_located((By.ID, 'TabContent'))
    WebDriverWait(driver, 10).until(element_present)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    time.sleep(2)  
    
    div_tags = soup.find_all('li', style ='line-height: 36px;margin-left: 2px;height: 36px')
    
    if not div_tags:
        print("沒有更多的內容可以讀取，跳出迴圈。")
        break 
    
    for div in div_tags:
        time.sleep(2)
        # 提取抬頭
        title_tag = div.find(class_="titleHiding ulAWidth")
        if title_tag:
            title = title_tag.text.strip()
            ws[f'A{n}'] = title
    
        # 提取發布日期
        time_date_tag = div.find("span", style="float: right;")
        if time_date_tag:
            date = time_date_tag.text.strip()
            ws[f'B{n}'] = date
            
        # 提取Url
        a_tag = div.find('a')
        if a_tag:
            href = a_tag.get('href')
            href_get.append(href)  
            ws[f'C{n}'] = 'https://www.zycg.gov.cn' + href  # 完整URL
            n += 1  # 增加行數
        print(f"抬頭: {title}\n時間: {date}")

    count_per_page = len(div_tags)

######翻頁邏輯######
    if count_per_page < max_per_page:
        print("已到達最後一頁，跳出迴圈。")
        break
    
    try:
        next_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="page"]/div/button[3]'))
        )
        is_disabled = next_button.get_attribute('style')
        if is_disabled == 'true':
            print("Next button is disabled, no further pages.")
            break
        
        next_button.click()
        time.sleep(2)  
    except Exception as e:
        print(f"Error navigating to the next page: {e}")
        break  

wb.save(excel_file)
wb.close()


################################讀取EXCEL執行####################################
import requests
from openpyxl import Workbook
from bs4 import BeautifulSoup
import time
import datetime
import re  # 用於正則表達式
from openpyxl import load_workbook

# 設定 Excel 檔案路徑
excel_file = "C:/Users/11020964/Desktop/py/爬蟲/大陸/中央政府採購網0113.xlsx"
wb = load_workbook(excel_file)
ws = wb.active 

href_get = []
row = 2
while ws[f'C{row}'].value:
    href_get.append(ws[f'C{row}'].value)
    row += 1  # 逐行讀取

# 打印網址列表以確認
# print(href_get)

# 開始計時
start_time = time.time()
current_date = datetime.datetime.now().strftime('%Y%m%d')

# 創建新的 Excel 檔案來保存結果
output_file = "output.xlsx"
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "Data"

# 定義匹配「一」到「九」的正則表達式
number_pattern = re.compile(r'^[一二三四五六七八九]、')
row = 1  # Excel 開始的列
# 開始爬取和寫入資料
for href in href_get:
    try:
        # 發送 HTTP 請求
        response = requests.get(href, timeout=10)
        response.raise_for_status()  # 確保請求成功

        # 設定編碼格式，修正文字顯示問題
        response.encoding = 'utf-8'  # 或嘗試 'gbk'
        
        soup = BeautifulSoup(response.text, 'html.parser')

        # 擷取所有符合條件的 <p> 標籤
        tags = soup.find('div',class_="WordSection1")
        
        if tags:
            content = ""
            current_column = 1  # 用來控制當前列位置
            for tag in tags:
                # 修正文字內容並去掉多餘空白
                tag_text = tag.text.strip()
                
                # 檢查是否以「一」到「九」開頭
                match = number_pattern.match(tag_text)
                if match:
                    # 如果有現有內容，寫入上一列
                    if content:
                        ws_out.cell(row=row, column=current_column, value=content.strip())
                        current_column += 1  # 移動到下一列
                    # 開始新段落
                    content = tag_text  # 當前段落
                else:
                    # 如果不是以「一」到「九」開頭，將內容添加到當前段落
                    content += " " + tag_text

            # 最後一段內容寫入
            if content:
                ws_out.cell(row=row, column=current_column, value=content.strip())

            row += 1  # 換到下一行

        else:
            # 擷取所有符合條件的 <p> 標籤
            tags = soup.find(id='printArea')
            if tags:
                content = ""
                current_column = 1  # 用來控制當前列位置
                for tag in tags:
                    # 修正文字內容並去掉多餘空白
                    tag_text = tag.text.strip()
                    
                    # 檢查是否以「一」到「九」開頭
                    match = number_pattern.match(tag_text)
                    if match:
                        # 如果有現有內容，寫入上一列
                        if content:
                            ws_out.cell(row=row, column=current_column, value=content.strip())
                            current_column += 1  # 移動到下一列
                        # 開始新段落
                        content = tag_text  # 當前段落
                    else:
                        # 如果不是以「一」到「九」開頭，將內容添加到當前段落
                        content += " " + tag_text

                # 最後一段內容寫入
                if content:
                    ws_out.cell(row=row, column=current_column, value=content.strip())

                row += 1  # 換到下一行


    except requests.exceptions.RequestException as e:
        print(f"Error occurred for {href}: {e}")
        ws_out[f"A{row}"] = f"Error: {e}"
        row += 1

# 保存 Excel 檔案
wb_out.save(output_file)
end_time = time.time()
print(f"Data saved to {output_file}")
print(f"Time taken: {end_time - start_time:.2f} seconds")
