from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import time
import openpyxl
import pandas as pd
from openpyxl import Workbook
from selenium.webdriver.common.keys import Keys


keyword = "Flooring"

# 瀏覽器設定
option = webdriver.ChromeOptions()
prefs = {"profile.default_content_setting_values": {"notifications": 2}}
option.add_experimental_option("prefs", prefs)

driver = webdriver.Chrome(service=Service(
    executable_path="C:\\Users\\Desktop\\python\\chromedriver.exe"), options=option)
base_url = 'https://www.hidubai.com/'
driver.get(base_url)


WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
    (By.CSS_SELECTOR, 'input[placeholder="Find food, spas, companies, suppliers, ..."]'))).send_keys(keyword)
WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
    (By.CSS_SELECTOR, 'button[class="btn btn-default searchButton"'))).click()
excel_file = f"C:/Users/Desktop/py/爬蟲/UAE/{keyword}.xlsx"

try:
    wb = openpyxl.load_workbook(excel_file)
except FileNotFoundError:
    wb = Workbook()  
    wb.save(excel_file)  
    ws = wb.active

wb = openpyxl.load_workbook(excel_file)
ws = wb.active

ws.title = f'{keyword}'
ws['A1'] = 'Company_name'
ws['B1'] = 'Maps'
ws['C1'] = 'Address'
ws['D1'] = 'Local number'
ws['E1'] = 'Email'
ws['F1'] = 'Url'
n = 2  

href_get = []  
max_per_page = 80 
count_per_page = 0

while True:
    element_present = EC.presence_of_element_located((By.CLASS_NAME, 'businessText'))
    WebDriverWait(driver, 10).until(element_present)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    time.sleep(2)  
    
    div_tags = soup.find_all('div', class_='col-xs-12 col-md-6 col-lg-4 elementContainer')
    
    if not div_tags:
        print("沒有更多的內容可以讀取，跳出迴圈。")
        break 

    for div in div_tags:
        a_tag = div.find('a')
        if a_tag:
            href = a_tag.get('href') 
            if href and href.startswith('/businesses'):
                href = base_url + href.lstrip('/')  
                href_get.append(href)  
                ws[f'F{n}'] = href 
                n += 1 
                print(href)

    count_per_page = len(div_tags)
    
######翻頁邏輯######
    if count_per_page < max_per_page:
        print("已到達最後一頁，跳出迴圈。")
        break
    
    try:
        next_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//a[@ng-click="selectPage(page + 1, $event)"]'))
        )
        is_disabled = next_button.get_attribute('ng-disabled')
        if is_disabled == 'true':
            print("Next button is disabled, no further pages.")
            break
        
        next_button.click()
        time.sleep(5)  
    except Exception as e:
        print(f"Error navigating to the next page: {e}")
        break  
wb.save(excel_file)


#################################讀取EXCEL執行####################################
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import openpyxl
import pandas as pd
import os
import datetime


option = webdriver.ChromeOptions()
option.add_argument('--headless')  # 啟用無頭模式


driver = webdriver.Chrome(service=Service(
    executable_path="C:\\Users\\Desktop\\python\\chromedriver.exe"), options=option)

try:
    wb = openpyxl.load_workbook(excel_file)
except FileNotFoundError:
    wb = Workbook()  
    wb.save(excel_file)  
    ws = wb.active

wb = openpyxl.load_workbook(excel_file)
ws = wb.active

ws.title = f'{keyword}'
ws['A1'] = 'Company_name'
ws['B1'] = 'Maps'
ws['C1'] = 'Address'
ws['D1'] = 'Local number'
ws['E1'] = 'Email'
ws['F1'] = 'Url'
n = 2  


href_get = []
row = 2  
while ws[f'F{row}'].value:
    href_get.append(ws[f'F{row}'].value)
    row += 1  

start_time = time.time()
current_date = datetime.datetime.now().strftime('%Y%m%d')

n = 2
for href in href_get:
    
    # 打開新的視窗
    driver.execute_script("window.open(arguments[0]);", href)
    WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(len(driver.window_handles)))
    driver.switch_to.window(driver.window_handles[-1])
    time.sleep(1)

    # 提取公司名稱
    soup = BeautifulSoup(driver.page_source, "html.parser")
    company_name_tag = soup.find("div", class_="bizTitleContainer")
    if company_name_tag:
        company = company_name_tag.text.strip()
        ws[f'A{n}'] = company
    else:
        ws[f'A{n}'] = "Company not found"
        
    # 提取maps
    maps_tag = soup.find("div", class_="bizLocation")
    if maps_tag:
        maps = maps_tag.text.strip()
        ws[f'B{n}'] = maps
    else:
        ws[f'B{n}'] = "maps not found"


    # 提取地址
    try:
        biz_map_div = soup.find("div", id="bizMap")
        if biz_map_div:
            address_tag = biz_map_div.find("h5", class_='ng-binding ng-scope')
            if address_tag:
                address = address_tag.find("span").text.strip()
            else:
                address = "Address not found"
        else:
            address = "bizMap div not found"
        ws[f'C{n}'] = address
    except Exception as e:
        ws[f'C{n}'] = "Address not found"


    # 提取市話（local number）
    try:
        phone_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//div[@title="Contact Phone"]')))
        phone_button.click()
        time.sleep(2)

        detail_soup = BeautifulSoup(driver.page_source, "html.parser")
        local_number = detail_soup.find('div', title="Contact Phone").find('span', class_='valueField visibleXs ng-binding').text.strip()
        ws[f'D{n}'] = local_number
    except Exception as e:
        ws[f'D{n}'] = "Local number not found"
        
    # 提取Email
    try:
        mail_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//div[@title="email"]')))
        mail_button.click()
        time.sleep(2)

        detail_soup = BeautifulSoup(driver.page_source, "html.parser")
        Email = detail_soup.find('a', title="email").find('span', class_='valueField hidden-xs ng-binding').text.strip()
        ws[f'E{n}'] = Email
    except Exception as e:
        ws[f'E{n}'] = "email not found"
        
    print(f'{n-1}... {company}: \
          \n地址: {address}\
          \nMaps: {maps}\
          \n市話: {local_number}')

    driver.close()
    driver.switch_to.window(driver.window_handles[0])
    n = n + 1
    
wb.save(excel_file)
end_time = time.time()
execution_time = end_time - start_time
hours, rem = divmod(execution_time, 3600)
minutes, seconds = divmod(rem, 60)
save_path = r'Z:\28_數據中心部\爬蟲名單\加盟商\UAE'
a = pd.read_excel(excel_file)
data = a.drop_duplicates(subset=['Company_name'], keep='first', inplace=False)

file_name = f'{current_date}_{keyword}_deduplication.xlsx'
full_path = os.path.join(save_path, file_name)
data.to_excel(full_path, index=False)

print(f'檔案已儲存至: {full_path}')
print(f'程式執行時間: {int(hours)} 小時 {int(minutes)} 分 {seconds:.2f} 秒')


