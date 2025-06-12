import requests
from bs4 import BeautifulSoup
import openpyxl
import json
import os
import time

'''
get page_namber
'''
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Interior Design"
ws['A1'] = 'Company_name'
ws['B1'] = 'Industry'
ws['C1'] = 'Address'
ws['D1'] = 'Phone'
ws['E1'] = 'Email'
ws['F1'] = 'Url'
ws['G1'] = 'update'

n = 2  

url = 'https://www.yellowpages.vn/cls/111010/do-go-noi-that-san-xuat-va-kinh-doanh-do-go-noi-that.html?page=1'


res = requests.get(url)


if res.status_code == 200:
    print("請求成功！")
else:
    print(f"請求失敗，狀態碼：{res.status_code}")


html = res.text

soup = BeautifulSoup(html, 'html.parser')

page_numbers = soup.find('div', {"class": 'mt-4 bg-info'}).find_all('a')
second_last_page = page_numbers[-2].text.strip()  

second_last_page_no_dot = second_last_page.replace('.', '')
second_last_page2 = int(second_last_page_no_dot)  
print(second_last_page2)  


'''
執行
'''


page = 1
second_last_page2 = 3

while page <= second_last_page2:
    print(f'目前第{page}頁')
    url = f'https://www.yellowpages.vn/cls/111010/do-go-noi-that-san-xuat-va-kinh-doanh-do-go-noi-that.html?page={page}'
    

    res = requests.get(url)
    soup = BeautifulSoup(res.text, 'html.parser')  
    

    divs = soup.find_all('div', {'class': 'yp_noidunglistings'})
    for div in divs:
        time.sleep(0.08)
        #公司
        company_name = div.find('h2' , {'class':'fs-5 pb-0 text-capitalize'})
        if company_name:
            ws[f'A{n}'] = company_name.text.strip()  
        else:
            ws[f'A{n}'] = 'Company not found'
            
        #Industry
        Industry_icon = div.find('i', {'class': 'fa fa-regular fa-clock pe-1'})
        if Industry_icon:
            Industry_id = Industry_icon.find_next('i')
            if Industry_id:
                ws[f'B{n}'] = Industry_id.text.strip()
            else:
                ws[f'B{n}'] = 'Industry type not found'
        else:
            ws[f'B{n}'] = 'Industry type not found'

            
        #地址
        address_tag = div.find('i' , {'class':'fa fa-solid fa-location-arrow text-black-50 pe-1'})
        if address_tag:
            ws[f'C{n}'] = address_tag.text.strip()  
        else:
            ws[f'C{n}'] = 'address not found'
            
            
        print('-' * 30)
        n += 1  # 下一行寫入數據
    page += 1  # 進入下一頁
    # time.sleep(10)

# 保存 Excel 檔案
wb.save('8_interior_design.xlsx')
print("資料已成功寫入")


