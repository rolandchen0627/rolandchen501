import requests
from bs4 import BeautifulSoup
import openpyxl
import json
import os
import time
from openpyxl import Workbook
import pandas as pd
import datetime

'''
get page_namber
'''
keyword = "Đồ gỗ Nội thất"
start_time = time.time()
current_date = datetime.datetime.now().strftime('%Y%m%d')
excel_file = f"C:/Users/11020964/Desktop/py/爬蟲/VN/{keyword}.xlsx"

try:
    wb = openpyxl.load_workbook(excel_file)
except FileNotFoundError:
    wb = Workbook()  
    wb.save(excel_file)
    ws = wb.active

wb = openpyxl.load_workbook(excel_file)
ws = wb.active
ws['A1'] = 'Company_name'
ws['B1'] = 'Address'
ws['C1'] = 'Phone'
ws['D1'] = 'Tax_code'
ws['E1'] = 'Short_name'
ws['F1'] = 'International name'
ws['G1'] = 'Management unit'
ws['H1'] = 'Date of establishment'
ws['I1'] = 'Business type'
ws['J1'] = 'State'

url = 'https://yp.vn/page/1/?s=%C4%90%E1%BB%93+g%E1%BB%97+N%E1%BB%99i+th%E1%BA%A5t'


def get_headers():    
    return  {
    "Host": "yp.vn",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:131.0) Gecko/20100101 Firefox/131.0",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/png,image/svg+xml,*/*;q=0.8",
    "Accept-Language": "zh-TW,zh;q=0.8,en-US;q=0.5,en;q=0.3",
    "Accept-Encoding": "gzip, deflate, br, zstd",
    "Referer": "https://yp.vn/",
    "Connection": "keep-alive",
    "Cookie": "_ga_MKBG9V1CD4=GS1.1.1729323857.8.0.1729323868.0.0.0; _ga=GA1.1.1962793211.1728020192; _ga_1NEM7EQ4FP=GS1.1.1729323856.7.0.1729323868.0.0.0; __zi=2000.SSZzejyD0SvbdEhmtLaVW6k4gApFKm6IAPQ-ky0D29DYsxozm0i8bcUGlBt1JmcJPPcrl30nDW.1; cf_clearance=Lt9uEHlqSChcd4B4KhpIQgkEYtCLwe4mXZv4oucksPg-1729323854-1.2.1.1-Bd9uCH.E6Y3MgkNnztdl.ASLbIRWYGfc8NMrsTH_CLXCJNYDX3.G6uUlvsitGNNeKyl_QD9yfnnd1hxqp4lhPnj_ZxiagZn9in2p.JRUE94VaF4ce5FDo89CwbH_CJ1FSXrJv5lLkvp.jRfrUjA3YlAtzFFW9mFjXDVsYd1iH71ibN5djFVw1Lt7FNf1byWcTnwNo.sOVESjimY3OSW5rAfS7.F81rYyZi.iKDUnT.qNm5uRT4Oc9w_XDLUWWh7pZhOpvRbltVq006qXUHewEE6bUBpKJo7yRXk5mTbXh5WSL75A55deLUMXrIoKlB_QHgp..lqyXkLyFlld3BE.k.74EW_lnScw2MGzsYv4S9pjhp.PrPnlEilZmg8lK0ocoLMLW27XP4QgwX9enbZLaw",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "same-origin",
    "Sec-Fetch-User": "?1",
    "Priority": "u=0, i"}

res = requests.get(url, headers=get_headers())

if res.status_code == 200:
    print("請求成功！")
else:
    print(f"請求失敗，狀態碼：{res.status_code}")


html = res.text

soup = BeautifulSoup(html, 'html.parser')

page_numbers = soup.find('ul', {"class": 'page-numbers'}).find_all('a')
second_last_page = page_numbers[-2].text.strip()  

second_last_page_no_dot = second_last_page.replace('.', '')
second_last_page2 = int(second_last_page_no_dot)  
print(second_last_page2)  


'''
執行
'''

n = 2
page = 1
last_page2 = 5
##################替換second_last_page2################
while page < last_page2:
    print(f'目前第{page}頁')
    url = f'https://yp.vn/page/{page}/?s=N%E1%BB%99i+th%E1%BA%A5t+v%C3%A0+X%C3%A2y+d%E1%BB%B1ng'
    

    res = requests.get(url, headers=get_headers())
    soup = BeautifulSoup(res.text, 'html.parser')  
    

    divs = soup.find_all('div', {'class': 'col-lg-12 col-md-12 pf-itempagedetail-element pf-tabfirst pfnewbglppage'})
    for div in divs:
        time.sleep(3)
        #公司
        company_name = div.find('h1')
        if company_name:
            ws[f'A{n}'] = company_name.text.strip()  
        else:
            ws[f'A{n}'] = 'Company not found'  
        
        #地址
        address_tag = div.find('span')
        if address_tag:
            ws[f'B{n}'] = address_tag.text.strip()  
        else:
            ws[f'B{n}'] = 'Address not found'  
        #電話
        phone_icon = div.find('i', {'class': 'fa fa-phone'})
        if phone_icon:
            phone_td = phone_icon.find_next('td')
            if phone_td:
                ws[f'C{n}'] = phone_td.text.strip()
            else:
                ws[f'C{n}'] = 'Phone not found'
        else:
            ws[f'C{n}'] = 'Phone not found'
        #稅碼
        tax_icon = div.find('i', {'class': 'fa fa-code'})
        if tax_icon:
            tax_td = tax_icon.find_next('td')
            if tax_td:
                ws[f'D{n}'] = tax_td.text.strip()
            else:
                ws[f'D{n}'] = 'tax not found'
        else:
            ws[f'D{n}'] = 'tax not found'
            
        #簡稱
        short_icon = div.find('i', {'class': 'fa fa-random'})
        if short_icon:
            short_td = short_icon.find_next('td')
            if short_td:
                ws[f'E{n}'] = short_td.text.strip()
            else:
                ws[f'E{n}'] = 'short name not found'
        else:
            ws[f'E{n}'] = 'short name not found'
        
        #國際名稱    
        inter_icon = div.find('i', {'class': 'fa fa-globe'})
        if inter_icon:
            inter_td = inter_icon.find_next('td')
            if inter_td:
                ws[f'F{n}'] = inter_td.text.strip()
            else:
                ws[f'F{n}'] = 'international name not found'
        else:
            ws[f'F{n}'] = 'international name not found'
            
        #管理單位
        mang_icon = div.find('i', {'class': 'fa fa-user'})
        if mang_icon:
            mang_td = mang_icon.find_next('td')
            if mang_td:
                ws[f'G{n}'] = mang_td.text.strip()
            else:
                ws[f'G{n}'] = 'management_unit not found'
        else:
            ws[f'G{n}'] = 'management_unit not found'
            
        #成立日期
        date_icon = div.find('i', {'class': 'fa fa-calendar'})
        if date_icon:
            date_td = date_icon.find_next('td')
            if date_td:
                ws[f'H{n}'] = date_td.text.strip()
            else:
                ws[f'H{n}'] = 'Date of establishment not found'
        else:
            ws[f'H{n}'] = 'Date of establishment not found'
            
        #業務類型
        busin_icon = div.find('i', {'class': 'fa fa-industry'})
        if busin_icon:
            busin_td = busin_icon.find_next('td')
            if busin_td:
                ws[f'I{n}'] = busin_td.text.strip()
            else:
                ws[f'I{n}'] = 'Business type not found'
        else:
            ws[f'I{n}'] = 'Business type not found'
            
        #狀態
        info_icon = div.find('i', {'class': 'fa fa-info-circle'})
        if info_icon:
            info_td = info_icon.find_next('td')
            if info_td:
                ws[f'J{n}'] = info_td.text.strip()
            else:
                ws[f'J{n}'] = 'info not found'
        else:
            ws[f'J{n}'] = 'info not found'
            
            
        print(f"""公司名稱: {ws[f'A{n}'].value} 地址: {ws[f'B{n}'].value} 電話: {ws[f'C{n}'].value} 成立日期: {ws[f'H{n}'].value} 狀態: {ws[f'J{n}'].value}""")
        print('-' * 30)
        n += 1  
    page += 1  
    time.sleep(10)

wb.save(excel_file)

end_time = time.time()
execution_time = end_time - start_time
hours, rem = divmod(execution_time, 3600)
minutes, seconds = divmod(rem, 60)
save_path = r'Z:\28_數據中心部\爬蟲名單\加盟商\VN'
a = pd.read_excel(excel_file)

# 檢查重複值並去重
data = a.drop_duplicates(subset=['Address'], keep='first', inplace=False)
file_name = f'{current_date}_{keyword}_deduplication.xlsx'
full_path = os.path.join(save_path, file_name)
data.to_excel(full_path, index=False)

print(f'檔案已儲存至: {full_path}')
print(f'程式執行時間: {int(hours)} 小時 {int(minutes)} 分 {seconds:.2f} 秒')



