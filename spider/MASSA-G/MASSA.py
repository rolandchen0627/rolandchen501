import os
import re
import time
import requests
import pandas as pd
from bs4 import BeautifulSoup
from datetime import datetime

from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter


def scrape_jobs(): 
    """開始爬取產品資訊並下載圖片（1~25頁）"""
    job_data = []
    keyword = "對鍊"
    today_date = datetime.today().strftime('%Y-%m-%d')

    output_folder = r'C:\Users\11020964.TPTWKD\Desktop\爬蟲\精品'
    image_folder = os.path.join(output_folder, "images")
    os.makedirs(output_folder, exist_ok=True)
    os.makedirs(image_folder, exist_ok=True)
    output_file = os.path.join(output_folder, f'MASSA-G_{today_date}_{keyword}.xlsx')

    base_url = 'https://www.massa-g.com.tw'
    global_idx = 1  # 用來連續編號所有產品

    for page in range(1, 3):  # 頁數 1 到 25
        url = f'{base_url}/PDList.asp?pp1=06&pp2=&pp3=&q=&ob=E&pageno={page}'
        res = requests.get(url)
        res.raise_for_status()
        soup = BeautifulSoup(res.text, 'html.parser')

        articles = soup.find_all('div', style="border:1px solid #cccccc; margin-bottom:10px;")

        for article in articles:
            # 名稱
            name_tag = article.find('span', style="color:#000000; font-weight:bold;")
            product_name = name_tag.get_text(strip=True) if name_tag else "未知產品"

            # 價格（在兄弟 table 中）
            price_table = article.find_next_sibling('table')
            if price_table:
                price_span = price_table.find('span', style=lambda s: s and 'font-size:18pt' in s and 'font-weight:bold' in s)
                product_price = price_span.get_text(strip=True) if price_span else "未知價格"
            else:
                product_price = "未知價格"

            # 圖片
            img_tag = article.find('img')
            img_url = base_url + img_tag['src'] if img_tag and img_tag.get('src') else ""

            # 下載圖片
            image_filename = "無圖片"
            if img_url:
                try:
                    sanitized_name = re.sub(r'[\\/:"*?<>|]', '_', product_name)
                    image_filename = os.path.join(image_folder, f'{global_idx:03d}_{sanitized_name}.jpg')
                    img_res = requests.get(img_url)
                    img_res.raise_for_status()
                    with open(image_filename, 'wb') as f:
                        f.write(img_res.content)
                except Exception as e:
                    print(f"⚠️ 圖片下載失敗: {img_url}：{e}")
                    image_filename = "下載失敗"

            job_info = {
                "產品種類": keyword,
                "產品名稱": product_name,
                "價格": product_price,
                "圖片網址": img_url,
                "本地圖片路徑": image_filename
            }
            job_data.append(job_info)

            print(f"{global_idx:03d}｜{product_name}｜{product_price}｜圖片：{'✅' if img_url else '❌'}")
            global_idx += 1

    # 匯出 Excel
    df = pd.DataFrame(job_data)
    df.to_excel(output_file, index=False)
    print(f"\n✅ 資料已匯出至：{output_file}")

    return job_data, output_file


# 測試執行
if __name__ == "__main__":
    scrape_jobs()
    


def save_to_excel_with_images(job_data, output_file):
    """儲存產品資訊到 Excel 並插入圖片"""
    wb = Workbook()
    ws = wb.active
    ws.title = "產品清單"

    headers = ["產品種類", "產品名稱", "價格", "圖片"]
    ws.append(headers)

    for idx, item in enumerate(job_data, start=2):  # 從第2列開始寫入資料
        ws.cell(row=idx, column=1, value=item["產品種類"])
        ws.cell(row=idx, column=2, value=item["產品名稱"])
        ws.cell(row=idx, column=3, value=item["價格"])

        # 插入圖片
        img_path = item.get("本地圖片路徑")
        if os.path.exists(img_path) and img_path.lower().endswith((".jpg", ".png")):
            try:
                img = OpenpyxlImage(img_path)
                img.width = 80  # 圖片寬
                img.height = 80  # 圖片高
                ws.row_dimensions[idx].height = 60  # Excel 行高，讓圖片顯示不被裁切
                ws.add_image(img, f"D{idx}")  # 插入圖片到 E 欄
            except Exception as e:
                print(f"⚠️ 插入圖片失敗 {img_path}：{e}")

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 25

    wb.save(output_file)


def main():
    job_data, output_file = scrape_jobs()
    save_to_excel_with_images(job_data, output_file)


if __name__ == "__main__":
    main()
