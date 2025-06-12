from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import time
import openpyxl
import pandas as pd
from openpyxl import Workbook
from selenium.webdriver.common.keys import Keys
import requests

# 瀏覽器設定
option = webdriver.ChromeOptions()
prefs = {"profile.default_content_setting_values": {"notifications": 2}}
option.add_experimental_option("prefs", prefs)

driver = webdriver.Chrome(service=Service(
    executable_path="C:\\Users\\11020964\\Desktop\\python\\chromedriver.exe"), options=option)
base_url = 'http://www.cebpubservice.com/ctpsp_iiss/searchbusinesstypebeforedooraction/getSearch.do#'
driver.get(base_url)


WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.ID, "11"))  # ID 不需要 "id="，只需值
).click()

WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.ID, "winBidTableId"))  # ID 不需要 "id="，只需值
).click()
excel_file = "C:/Users/11020964/Desktop/py/爬蟲/大陸/中央招標.xlsx"



    # 提取公司名稱
    soup = BeautifulSoup(driver.page_source, "html.parser")
    company_name_tag = soup.find("div", class_="bizTitleContainer")
    if company_name_tag:
        company = company_name_tag.text.strip()
        ws[f'A{n}'] = company
    else:
        ws[f'A{n}'] = "Company not found"
        
    # 提取maps
    maps_tag = soup.find("div", class_="bizLocation")
    if maps_tag:
        maps = maps_tag.text.strip()
        ws[f'B{n}'] = maps
    else:
        ws[f'B{n}'] = "maps not found"


    # 提取地址
    try:
        biz_map_div = soup.find("div", id="bizMap")
        if biz_map_div:
            address_tag = biz_map_div.find("h5", class_='ng-binding ng-scope')
            if address_tag:
                address = address_tag.find("span").text.strip()
            else:
                address = "Address not found"
        else:
            address = "bizMap div not found"
        ws[f'C{n}'] = address
    except Exception as e:
        ws[f'C{n}'] = "Address not found"


    # 提取市話（local number）
    try:
        phone_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//div[@title="Contact Phone"]')))
        phone_button.click()
        time.sleep(2)

        detail_soup = BeautifulSoup(driver.page_source, "html.parser")
        local_number = detail_soup.find('div', title="Contact Phone").find('span', class_='valueField visibleXs ng-binding').text.strip()
        ws[f'D{n}'] = local_number
    except Exception as e:
        ws[f'D{n}'] = "Local number not found"
        
    # 提取Email
    try:
        mail_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//div[@title="email"]')))
        mail_button.click()
        time.sleep(2)

        detail_soup = BeautifulSoup(driver.page_source, "html.parser")
        Email = detail_soup.find('a', title="email").find('span', class_='valueField hidden-xs ng-binding').text.strip()
        ws[f'E{n}'] = Email
    except Exception as e:
        ws[f'E{n}'] = "email not found"
        
    print(f'{n-1}... {company}: \
          \n地址: {address}\
          \nMaps: {maps}\
          \n市話: {local_number}')

    driver.close()
    driver.switch_to.window(driver.window_handles[0])
    n = n + 1
    
wb.save(excel_file)
end_time = time.time()
execution_time = end_time - start_time
hours, rem = divmod(execution_time, 3600)
minutes, seconds = divmod(rem, 60)
save_path = r'Z:\28_數據中心部\爬蟲名單\加盟商\UAE'
a = pd.read_excel(excel_file)
data = a.drop_duplicates(subset=['Company_name'], keep='first', inplace=False)

file_name = f'{current_date}_{keyword}_deduplication.xlsx'
full_path = os.path.join(save_path, file_name)
data.to_excel(full_path, index=False)

print(f'檔案已儲存至: {full_path}')
print(f'程式執行時間: {int(hours)} 小時 {int(minutes)} 分 {seconds:.2f} 秒')
